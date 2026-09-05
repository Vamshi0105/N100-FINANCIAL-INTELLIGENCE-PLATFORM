from __future__ import annotations

import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]

DEFAULT_DB_PATH = (
    PROJECT_ROOT
    / "data"
    / "nifty100.db"
)

DEFAULT_OUTPUT_PATH = (
    PROJECT_ROOT
    / "output"
    / "peer_comparison.xlsx"
)


# -------------------------------------------------
# Day 18 metric names mapped to financial_ratios
# -------------------------------------------------

METRICS = [
    {
        "label": "ROE",
        "db_column": "return_on_equity_pct",
        "percentile_metric": "ROE",
    },
    {
        "label": "ROCE",
        "db_column": "return_on_capital_employed_pct",
        "percentile_metric": "ROCE",
    },
    {
        "label": "Net Profit Margin",
        "db_column": "net_profit_margin_pct",
        "percentile_metric": "Net Profit Margin",
    },
    {
        "label": "D/E",
        "db_column": "debt_to_equity",
        "percentile_metric": "D/E",
    },
    {
        "label": "FCF",
        "db_column": "free_cash_flow_cr",
        "percentile_metric": "FCF",
    },
    {
        "label": "PAT CAGR 5yr",
        "db_column": "pat_cagr_5yr",
        "percentile_metric": "PAT CAGR 5yr",
    },
    {
        "label": "Revenue CAGR 5yr",
        "db_column": "revenue_cagr_5yr",
        "percentile_metric": "Revenue CAGR 5yr",
    },
    {
        "label": "EPS CAGR 5yr",
        "db_column": "eps_cagr_5yr",
        "percentile_metric": "EPS CAGR 5yr",
    },
    {
        "label": "Interest Coverage",
        "db_column": "interest_coverage",
        "percentile_metric": "Interest Coverage",
    },
    {
        "label": "Asset Turnover",
        "db_column": "asset_turnover",
        "percentile_metric": "Asset Turnover",
    },
]


# -------------------------------------------------
# Excel styles
# -------------------------------------------------

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

BENCHMARK_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFD966",
)

GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE",
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C",
)

RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)

SUMMARY_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAD3",
)

BOLD_FONT = Font(
    bold=True,
)


# -------------------------------------------------
# Database helpers
# -------------------------------------------------

def get_latest_annual_year(
    connection: sqlite3.Connection,
) -> str:
    """
    Get latest annual financial year.

    Example:
        2024-03
    """

    row = connection.execute(
        """
        SELECT MAX(year)
        FROM financial_ratios
        WHERE year LIKE '%-03'
        """
    ).fetchone()

    if row is None or row[0] is None:
        raise ValueError(
            "No annual financial ratio data found."
        )

    return str(row[0])


def load_peer_groups(
    connection: sqlite3.Connection,
) -> pd.DataFrame:
    """
    Load peer group assignments.
    """

    query = """
        SELECT
            peer_group_name,
            company_id,
            is_benchmark
        FROM peer_groups
        ORDER BY
            peer_group_name,
            company_id
    """

    df = pd.read_sql_query(
        query,
        connection,
    )

    if df.empty:
        return df

    df["company_id"] = (
        df["company_id"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    df["is_benchmark"] = (
        df["is_benchmark"]
        .fillna(0)
        .astype(int)
    )

    return df


def load_company_names(
    connection: sqlite3.Connection,
) -> pd.DataFrame:
    """
    Load company IDs and company names.
    """

    query = """
        SELECT
            id AS company_id,
            company_name
        FROM companies
    """

    df = pd.read_sql_query(
        query,
        connection,
    )

    if df.empty:
        return df

    df["company_id"] = (
        df["company_id"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    return df


def load_financial_metrics(
    connection: sqlite3.Connection,
    year: str,
) -> pd.DataFrame:
    """
    Load the 10 financial metrics
    for the selected annual year.
    """

    metric_columns = ",\n".join(
        metric["db_column"]
        for metric in METRICS
    )

    query = f"""
        SELECT
            company_id,
            year,
            {metric_columns}
        FROM financial_ratios
        WHERE year = ?
    """

    df = pd.read_sql_query(
        query,
        connection,
        params=(year,),
    )

    if df.empty:
        return df

    df["company_id"] = (
        df["company_id"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    return df


def load_peer_percentiles(
    connection: sqlite3.Connection,
    year: str,
) -> pd.DataFrame:
    """
    Load Day 18 peer percentile rankings.
    """

    query = """
        SELECT
            company_id,
            peer_group_name,
            metric,
            value,
            percentile_rank,
            year
        FROM peer_percentiles
        WHERE year = ?
    """

    df = pd.read_sql_query(
        query,
        connection,
        params=(year,),
    )

    if df.empty:
        return df

    df["company_id"] = (
        df["company_id"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    return df


# -------------------------------------------------
# Data preparation
# -------------------------------------------------

def build_peer_group_dataframe(
    peer_group_name: str,
    peer_groups_df: pd.DataFrame,
    company_names_df: pd.DataFrame,
    financial_df: pd.DataFrame,
    percentiles_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Build one complete DataFrame
    for one peer group.
    """

    group_df = peer_groups_df[
        peer_groups_df["peer_group_name"]
        == peer_group_name
    ].copy()

    # Add company names.
    group_df = group_df.merge(
        company_names_df,
        on="company_id",
        how="left",
    )

    # Add raw financial metrics.
    group_df = group_df.merge(
        financial_df.drop(
            columns=["year"],
            errors="ignore",
        ),
        on="company_id",
        how="left",
    )

    # -------------------------------------------------
    # Add percentile columns one metric at a time.
    # -------------------------------------------------

    for metric in METRICS:

        metric_label = metric[
            "percentile_metric"
        ]

        percentile_column = (
            f"{metric['label']} Percentile"
        )

        metric_percentiles = (
            percentiles_df[
                (
                    percentiles_df["peer_group_name"]
                    == peer_group_name
                )
                &
                (
                    percentiles_df["metric"]
                    == metric_label
                )
            ][
                [
                    "company_id",
                    "percentile_rank",
                ]
            ]
            .copy()
        )

        metric_percentiles = (
            metric_percentiles.rename(
                columns={
                    "percentile_rank":
                        percentile_column
                }
            )
        )

        group_df = group_df.merge(
            metric_percentiles,
            on="company_id",
            how="left",
        )

    # -------------------------------------------------
    # Rename raw database columns
    # to user-friendly Excel names.
    # -------------------------------------------------

    rename_map = {}

    for metric in METRICS:

        rename_map[
            metric["db_column"]
        ] = metric["label"]

    group_df = group_df.rename(
        columns=rename_map
    )

    # -------------------------------------------------
    # Final column order.
    # -------------------------------------------------

    columns = [
        "company_id",
        "company_name",
        "is_benchmark",
    ]

    for metric in METRICS:

        columns.append(
            metric["label"]
        )

        columns.append(
            f"{metric['label']} Percentile"
        )

    group_df = group_df[
        [
            column
            for column in columns
            if column in group_df.columns
        ]
    ]

    # Put benchmark company first.
    group_df = group_df.sort_values(
        by=[
            "is_benchmark",
            "company_id",
        ],
        ascending=[
            False,
            True,
        ],
    )

    return group_df.reset_index(
        drop=True
    )


# -------------------------------------------------
# Excel formatting
# -------------------------------------------------

def write_peer_group_sheet(
    workbook: Workbook,
    peer_group_name: str,
    group_df: pd.DataFrame,
    year: str,
) -> None:
    """
    Write one peer group to one Excel worksheet.
    """

    sheet_name = peer_group_name[:31]

    worksheet = workbook.create_sheet(
        title=sheet_name
    )

    # -------------------------------------------------
    # Title
    # -------------------------------------------------

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(group_df.columns) - 1,
    )

    title_cell = worksheet.cell(
        row=1,
        column=1,
    )

    title_cell.value = (
        f"{peer_group_name} "
        f"Peer Comparison — {year}"
    )

    title_cell.font = Font(
        bold=True,
        size=14,
        color="FFFFFF",
    )

    title_cell.fill = HEADER_FILL

    title_cell.alignment = Alignment(
        horizontal="center"
    )

    # -------------------------------------------------
    # Header row
    # -------------------------------------------------

    excel_columns = [
        column
        for column in group_df.columns
        if column != "is_benchmark"
    ]

    header_row = 3

    for column_number, column_name in enumerate(
        excel_columns,
        start=1,
    ):

        cell = worksheet.cell(
            row=header_row,
            column=column_number,
            value=column_name,
        )

        cell.fill = HEADER_FILL

        cell.font = HEADER_FONT

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    # -------------------------------------------------
    # Data rows
    # -------------------------------------------------

    data_start_row = header_row + 1

    percentile_columns = {
        f"{metric['label']} Percentile"
        for metric in METRICS
    }

    metric_columns = {
        metric["label"]
        for metric in METRICS
    }

    for dataframe_index, dataframe_row in (
        group_df.iterrows()
    ):

        excel_row = (
            data_start_row
            + dataframe_index
        )

        is_benchmark = (
            dataframe_row["is_benchmark"]
            == 1
        )

        excel_column = 1

        for column_name in excel_columns:

            value = dataframe_row[
                column_name
            ]

            if pd.isna(value):
                value = None

            cell = worksheet.cell(
                row=excel_row,
                column=excel_column,
                value=value,
            )

            cell.alignment = Alignment(
                vertical="center"
            )

            # -------------------------------------------------
            # Benchmark highlighting.
            #
            # Percentile cells keep their traffic-light colour
            # so both benchmark and ranking information remain
            # visually useful.
            # -------------------------------------------------

            if (
                is_benchmark
                and column_name
                not in percentile_columns
            ):

                cell.fill = BENCHMARK_FILL

                cell.font = BOLD_FONT

            # -------------------------------------------------
            # Percentile colour coding.
            # -------------------------------------------------

            if (
                column_name
                in percentile_columns
                and value is not None
            ):

                percentile_value = float(
                    value
                )

                if percentile_value >= 0.75:

                    cell.fill = GREEN_FILL

                elif percentile_value <= 0.25:

                    cell.fill = RED_FILL

                else:

                    cell.fill = YELLOW_FILL

                cell.number_format = (
                    "0.0%"
                )

            # -------------------------------------------------
            # Metric number formatting.
            # -------------------------------------------------

            if column_name in metric_columns:

                cell.number_format = (
                    "0.00"
                )

            excel_column += 1

    data_end_row = (
        data_start_row
        + len(group_df)
        - 1
    )

    # -------------------------------------------------
    # Median summary row.
    # -------------------------------------------------

    summary_row = (
        data_end_row
        + 2
    )

    worksheet.cell(
        row=summary_row,
        column=1,
        value="Peer Group Median",
    )

    worksheet.cell(
        row=summary_row,
        column=1,
    ).font = BOLD_FONT

    # Company name column gets blank.
    worksheet.cell(
        row=summary_row,
        column=2,
        value="",
    )

    for column_number, column_name in enumerate(
        excel_columns,
        start=1,
    ):

        cell = worksheet.cell(
            row=summary_row,
            column=column_number,
        )

        cell.fill = SUMMARY_FILL

        cell.font = BOLD_FONT

        # Median applies only to
        # raw financial metric columns.
        if column_name in metric_columns:

            column_letter = (
                get_column_letter(
                    column_number
                )
            )

            cell.value = (
                f"=MEDIAN("
                f"{column_letter}"
                f"{data_start_row}:"
                f"{column_letter}"
                f"{data_end_row}"
                f")"
            )

            cell.number_format = (
                "0.00"
            )

    # -------------------------------------------------
    # Freeze panes and filters.
    # -------------------------------------------------

    worksheet.freeze_panes = "C4"

    worksheet.auto_filter.ref = (
        f"A{header_row}:"
        f"{get_column_letter(len(excel_columns))}"
        f"{data_end_row}"
    )

    # -------------------------------------------------
    # Column widths.
    # -------------------------------------------------

    for column_number, column_name in enumerate(
        excel_columns,
        start=1,
    ):

        column_letter = (
            get_column_letter(
                column_number
            )
        )

        if column_name == "company_id":

            width = 16

        elif column_name == "company_name":

            width = 28

        elif "Percentile" in column_name:

            width = 16

        else:

            width = 18

        worksheet.column_dimensions[
            column_letter
        ].width = width

    # Row heights.
    worksheet.row_dimensions[
        header_row
    ].height = 32

    worksheet.row_dimensions[
        1
    ].height = 24


# -------------------------------------------------
# Main workflow
# -------------------------------------------------

def generate_peer_comparison_report(
    db_path: str | Path = DEFAULT_DB_PATH,
    output_path: str | Path = DEFAULT_OUTPUT_PATH,
) -> Path:
    """
    Generate the Day 20 peer comparison Excel report.
    """

    db_path = Path(
        db_path
    )

    output_path = Path(
        output_path
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with sqlite3.connect(
        db_path
    ) as connection:

        year = get_latest_annual_year(
            connection
        )

        peer_groups_df = (
            load_peer_groups(
                connection
            )
        )

        company_names_df = (
            load_company_names(
                connection
            )
        )

        financial_df = (
            load_financial_metrics(
                connection,
                year,
            )
        )

        percentiles_df = (
            load_peer_percentiles(
                connection,
                year,
            )
        )

    if peer_groups_df.empty:

        raise ValueError(
            "No peer groups found."
        )

    if percentiles_df.empty:

        raise ValueError(
            "No peer percentile rankings found. "
            "Run Day 18 first."
        )

    workbook = Workbook()

    # Remove default sheet.
    default_sheet = workbook.active

    workbook.remove(
        default_sheet
    )

    peer_group_names = sorted(
        peer_groups_df[
            "peer_group_name"
        ]
        .dropna()
        .unique()
        .tolist()
    )

    for peer_group_name in peer_group_names:

        group_df = (
            build_peer_group_dataframe(
                peer_group_name=
                    peer_group_name,
                peer_groups_df=
                    peer_groups_df,
                company_names_df=
                    company_names_df,
                financial_df=
                    financial_df,
                percentiles_df=
                    percentiles_df,
            )
        )

        write_peer_group_sheet(
            workbook=workbook,
            peer_group_name=
                peer_group_name,
            group_df=group_df,
            year=year,
        )

    workbook.save(
        output_path
    )

    print(
        "Peer comparison report created:"
    )

    print(
        output_path
    )

    print()

    print(
        "Peer groups:",
        len(peer_group_names),
    )

    print(
        "Reporting year:",
        year,
    )

    return output_path


if __name__ == "__main__":

    generate_peer_comparison_report()