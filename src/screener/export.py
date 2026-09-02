from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_PATH = Path("output/screener_output.xlsx")


# ---------------------------------------------------------
# KPI COLUMNS
# ---------------------------------------------------------

KPI_COLUMNS = [
    "company_id",
    "year",
    "broad_sector",
    "composite_quality_score",

    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",

    "free_cash_flow_cr",
    "fcf_cagr_5yr",
    "cfo_quality_ratio",

    "revenue_cagr_3yr",
    "revenue_cagr_5yr",

    "pat_cagr_3yr",
    "pat_cagr_5yr",

    "debt_to_equity",
    "interest_coverage",

    "sales",
    "net_profit",

    "pe_ratio",
    "pb_ratio",

    "dividend_yield_pct",
]


# ---------------------------------------------------------
# PRESET THRESHOLD MAPPING
# ---------------------------------------------------------

FILTER_TO_COLUMN = {

    "return_on_equity_pct_min":
        ("return_on_equity_pct", "min"),

    "debt_to_equity_max":
        ("debt_to_equity", "max"),

    "free_cash_flow_cr_min":
        ("free_cash_flow_cr", "min"),

    "revenue_cagr_3yr_min":
        ("revenue_cagr_3yr", "min"),

    "revenue_cagr_5yr_min":
        ("revenue_cagr_5yr", "min"),

    "pat_cagr_5yr_min":
        ("pat_cagr_5yr", "min"),

    "sales_min":
        ("sales", "min"),

    "pe_ratio_max":
        ("pe_ratio", "max"),

    "pb_ratio_max":
        ("pb_ratio", "max"),

    "dividend_yield_pct_min":
        ("dividend_yield_pct", "min"),
}


GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE",
)

RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)


# ---------------------------------------------------------
# EXPORT
# ---------------------------------------------------------

def export_screener_results(
    preset_results: dict,
    preset_filters: dict,
    output_path=OUTPUT_PATH,
):
    """
    Export screener results to Excel.

    One worksheet is created for each preset.
    """

    output_path = Path(output_path)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:

        for preset_name, dataframe in (
            preset_results.items()
        ):

            available_columns = [
                column
                for column in KPI_COLUMNS
                if column in dataframe.columns
            ]

            export_dataframe = dataframe[
                available_columns
            ].copy()

            export_dataframe = (
                export_dataframe
                .sort_values(
                    "composite_quality_score",
                    ascending=False,
                )
            )

            sheet_name = (
                preset_name
                .replace("_", " ")
                .title()
            )

            export_dataframe.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

    apply_excel_formatting(
        output_path,
        preset_filters,
    )

    return output_path


# ---------------------------------------------------------
# EXCEL FORMATTING
# ---------------------------------------------------------

def apply_excel_formatting(
    output_path,
    preset_filters,
):
    """
    Apply threshold-based green/red formatting.
    """

    workbook = load_workbook(
        output_path
    )

    for preset_name, filters in (
        preset_filters.items()
    ):

        sheet_name = (
            preset_name
            .replace("_", " ")
            .title()
        )

        if sheet_name not in workbook.sheetnames:
            continue

        worksheet = workbook[
            sheet_name
        ]

        # Map header name to Excel column number
        headers = {}

        for column_number, cell in enumerate(
            worksheet[1],
            start=1,
        ):

            headers[cell.value] = (
                column_number
            )

        for filter_name, threshold in (
            filters.items()
        ):

            if filter_name not in FILTER_TO_COLUMN:
                continue

            column_name, rule = (
                FILTER_TO_COLUMN[
                    filter_name
                ]
            )

            if column_name not in headers:
                continue

            column_number = headers[
                column_name
            ]

            for row_number in range(
                2,
                worksheet.max_row + 1,
            ):

                cell = worksheet.cell(
                    row=row_number,
                    column=column_number,
                )

                value = cell.value

                if value is None:
                    continue

                try:

                    value = float(value)

                except (
                    ValueError,
                    TypeError,
                ):

                    continue

                passes = False

                if rule == "min":
                    passes = (
                        value >= threshold
                    )

                elif rule == "max":
                    passes = (
                        value <= threshold
                    )

                cell.fill = (
                    GREEN_FILL
                    if passes
                    else RED_FILL
                )

        # Freeze header
        worksheet.freeze_panes = "A2"

        # Auto-size columns
        for column_cells in (
            worksheet.columns
        ):

            max_length = 0

            column_letter = (
                get_column_letter(
                    column_cells[0].column
                )
            )

            for cell in column_cells:

                if cell.value is None:
                    continue

                max_length = max(
                    max_length,
                    len(
                        str(
                            cell.value
                        )
                    ),
                )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                30,
            )

    workbook.save(
        output_path
    )
    